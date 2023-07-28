from http.client import FOUND
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk

def newinterface():
    #New Interface
    NewRoot = Toplevel()
    NewRoot.geometry("1030x780")
    NewRoot.title("New Interface")
    NewRoot.configure(bg="black")

    #Excel File
    excel_con = Workbook()
    excel_con = load_workbook("Sample_file.xlsx")
    excel_data = excel_con['Data']

    #ADD RECORD BBUTTON FUNCTION
    def add_record():
            student_no = studentno_Entry.get()
            full_name = fullname_Entry.get()
            email = email_Entry.get()
            gender = gender_var.get()
            course = course_var.get()
            contact_no = no_Entry.get()
            address = address_Entry.get()
            
            Found = False
            for each_cell in range(2, excel_data.max_row + 1):
                if student_no == excel_data["A"+ str(each_cell)].value or full_name == excel_data["B"+ str(each_cell)].value:
                    Found = True
                    break
            if Found == True:
                messagebox.showerror("Data", "Data Already Exist")
            else:
                lastrow = str(excel_data.max_row + 1)
                excel_data["A"+lastrow] = student_no
                excel_data["B"+lastrow] = full_name
                excel_data["C"+lastrow] = email
                excel_data["D"+lastrow] = gender
                excel_data["E"+lastrow] = course
                excel_data["F"+lastrow] = contact_no
                excel_data["G"+lastrow] = address

                excel_con.save("Sample_file.xlsx")
                messagebox.showinfo("Save Records", "Records saved successfully!")
                refresh_data(tv1)

            record = f"Student Ref: {student_no}\n"
            record += f"Full Name: {full_name}\n"
            record += f"Email: {email}\n"
            record += f"Gender: {gender}\n"
            record += f"Course: {course}\n"
            record += f"Contact No.: {contact_no}\n"
            record += f"Address: {address}\n"
            
            student_detailsT.insert(END, record)

            studentno_Entry.delete(0, END)
            fullname_Entry.delete(0, END)
            email_Entry.delete(0, END)
            gender_var.set("none")
            course_var.set("")
            no_Entry.delete(0, END)
            address_Entry.delete(0, END)
    #REFRESH FUNCTION
    def refresh_data(tree):
            tree.delete(*tree.get_children())
            data = get_updated_data()
            for each_cell in range(2, (excel_data.max_row)+1):
                tv1.insert(parent='', index="end", text=str(each_cell),values=(excel_data['A'+str(each_cell)].value,excel_data['B'+str(each_cell)].value, excel_data['C'+str(each_cell)].value, excel_data['D'+str(each_cell)].value, excel_data['E'+str(each_cell)].value, excel_data['F'+str(each_cell)].value, excel_data['G'+str(each_cell)].value, excel_data['H'+str(each_cell)].value))

    #UPDATE FUNCTION
    def get_updated_data():
            updated_value = list()
            for each_cell in range(2, (excel_data.max_row)+1):     
                updated_value.append([excel_data['A'+str(each_cell)].value,excel_data['B'+str(each_cell)].value, excel_data['C'+str(each_cell)].value, excel_data['D'+str(each_cell)].value,excel_data['E'+str(each_cell)].value,excel_data['F'+str(each_cell)].value,excel_data['G'+str(each_cell)].value,excel_data['H'+str(each_cell)].value])
            return updated_value

    #EXIT BUTTON FUNCTION
    def exit_interface():
        NewRoot.destroy()

    #PRINT BUTTON FUNCTION
    def print_records():
        student_details_text = student_detailsT.get("1.0", END)
        messagebox.showinfo("Print Records", student_details_text)
        student_details_text = student_detailsT.delete("1.0", END)

    #RESET BUTTON FUNCTION
    def reset_fields():
            studentno_Entry.delete(0, END)
            fullname_Entry.delete(0, END)
            email_Entry.delete(0, END)
            gender_var.set("none")
            course_var.set("")
            no_Entry.delete(0, END)
            address_Entry.delete(0, END)
            search_box.delete(0, END)
            student_detailsT.delete("1.0", END)
            refresh_data(tv1)

    def search_data():
        data = []

        for i in excel_data.iter_rows(values_only=True):
            data.append(i)

            tv1.delete(*tv1.get_children())
            lists = []
            for i in data:
                x = False
                for j in i:
                    if j != None:
                        if search_box.get().lower() in j.lower():
                            x = True
                            break
                if x:
                    lists.append(i)
            
            for i in lists:
                tv1.insert('', index=END, values=i)

    def delete_data():
        student_no = search_box.get()
        pos = 1
        Found = False
        for each_cell in excel_data.iter_rows(values_only=True):
            if student_no == each_cell[0]:
                Found = True
                break
            pos += 1
        if(Found == True):
            excel_data.delete_rows(pos)
            messagebox.showinfo("INFO","DATA DELETED")
            clear_entries()
        excel_con.save('Sample_file.xlsx')
        refresh_data(tv1)

    def clear_entries():
        studentno_Entry.delete(0, END)
        fullname_Entry.delete(0, END)
        email_Entry.delete(0, END)
        gender_var.set("none")
        course_var.set("")
        no_Entry.delete(0, END)
        search_box.delete(0, END)
        address_Entry.delete(0, END)

    def edit_data():
        search_bo = search_box.get()
        for each_cell in range(2, (excel_data.max_row)+1):
            if search_bo ==  excel_data['A'+str(each_cell)].value:
                Found = True
                break
            else:
                Found=False
        if(Found == True):
            Edit_form = Toplevel()
            Edit_form.geometry('350x800')
            Edit_form.title('Edit Data from Excel')

            path = "edit.jpg"
            bg = Image.open(path)
            resize_bg = bg.resize((350, 800))
            bg = ImageTk.PhotoImage(resize_bg)
            my_canva = Canvas(Edit_form, height=800, width=350)
            my_canva.pack(fill="both")
            my_canva.create_image(0,0,image=bg, anchor="nw")

            
            student_noExcel = StringVar()
            full_nameExcel = StringVar()
            emailExcel = StringVar()
            genderExcel = StringVar()
            courseExcel = StringVar()
            noExcel = StringVar()
            addressExcel = StringVar()

            student_noTxt=Entry(Edit_form, width=18, font=('Helvetica',15),textvariable=student_noExcel, bg="gray")
            student_noTxt.pack()
            student_noTxt_w = my_canva.create_window(120,105, anchor="nw", window=student_noTxt)

            student_noChoice = IntVar()
            student_noChk = Checkbutton(Edit_form, text="same as before", variable=student_noChoice, command=lambda:get_existing_student_no(), bg="white")
            student_noChk.pack()
            student_noChk_w = my_canva.create_window(155,135, anchor="nw", window=student_noChk)
        

            full_nameTxt=Entry(Edit_form, width=18, font=('Helvetica',15),textvariable=full_nameExcel, bg="gray")
            full_nameTxt.pack()
            full_name_w = my_canva.create_window(120,190, anchor="nw", window=full_nameTxt)

            full_nameChoice = IntVar()
            full_nameChk = Checkbutton(Edit_form, text="same as before", variable=full_nameChoice, command=lambda:get_existing_full_name(), bg="white")
            full_nameChk.pack()
            full_nameChk_w = my_canva.create_window(155,220, anchor="nw", window=full_nameChk)


            emailTxt=Entry(Edit_form, width=18, font=('Helvetica',15),textvariable=emailExcel, bg="gray")
            emailTxt.pack()
            emailTxt_w = my_canva.create_window(120,270, anchor="nw", window=emailTxt)

            emailChoice = IntVar()
            emailChk = Checkbutton(Edit_form, text="same as before", variable=emailChoice, command=lambda:get_existing_email(), bg="white")
            emailChk.pack()
            emailChk_w = my_canva.create_window(155,300, anchor="nw", window=emailChk)


            genderTxt=Entry(Edit_form, width=18, font=('Helvetica',15),textvariable=genderExcel, bg="gray")
            genderTxt.pack()
            genderTxt_w = my_canva.create_window(120,350, anchor="nw", window=genderTxt)

            genderChoice = IntVar()
            genderChk = Checkbutton(Edit_form, text="same as before", variable=genderChoice, command=lambda:get_existing_gender(), bg="white")
            genderChk.pack()
            genderChk_w = my_canva.create_window(155,380, anchor="nw", window=genderChk)


            courseTxt=Entry(Edit_form, width=18, font=('Helvetica',15),textvariable=courseExcel, bg="gray")
            courseTxt.pack()
            courseTxt_w = my_canva.create_window(120,425, anchor="nw", window=courseTxt)

            courseChoice = IntVar()
            courseChk = Checkbutton(Edit_form, text="same as before", variable=courseChoice, command=lambda:get_existing_course(), bg="white")
            courseChk.pack()
            courseChk_w = my_canva.create_window(155,455, anchor="nw", window=courseChk)


            noTxt=Entry(Edit_form, width=18, font=('Helvetica',15),textvariable=noExcel, bg="gray")
            noTxt.pack()
            noTxt_w = my_canva.create_window(120,505, anchor="nw", window=noTxt)

            noChoice = IntVar()
            noChk = Checkbutton(Edit_form, text="same as before", variable=noChoice, command=lambda:get_existing_no(), bg="white")
            noChk.pack()
            noChk_w = my_canva.create_window(155,535, anchor="nw", window=noChk)


            addressTxt=Entry(Edit_form, width=18, font=('Helvetica',15),textvariable=addressExcel, bg="gray")
            addressTxt.pack()
            addressTxt_w = my_canva.create_window(120,580, anchor="nw", window=addressTxt)

            addressChoice = IntVar()
            addressChk = Checkbutton(Edit_form, text="same as before", variable=addressChoice, command=lambda:get_existing_address(), bg="white")
            addressChk.pack()
            addressChk_w = my_canva.create_window(155,615, anchor="nw", window=addressChk)

            def get_existing_student_no():
                if student_noChoice.get()==1:
                    student_noOld = excel_data['A'+str(each_cell)].value
                    student_noExcel.set(student_noOld)
                elif student_noChoice.get() ==0:
                    student_noExcel.set("")
            def get_existing_full_name():
                if full_nameChoice.get()==1:
                    full_nameOld = excel_data['B'+str(each_cell)].value
                    full_nameExcel.set(full_nameOld)
                elif full_nameChoice.get() ==0:
                    full_nameExcel.set("")
            def get_existing_email():
                if emailChoice.get()==1:
                    emailOld = excel_data['C'+str(each_cell)].value
                    emailExcel.set(emailOld)
                elif emailChoice.get() ==0:
                    emailExcel.set("")
            def get_existing_gender():
                if genderChoice.get()==1:
                    genderOld = excel_data['D'+str(each_cell)].value
                    genderExcel.set(genderOld)
                elif genderChoice.get() ==0:
                    genderExcel.set("")
            def get_existing_course():
                if courseChoice.get()==1:
                    courseOld = excel_data['E'+str(each_cell)].value
                    courseExcel.set(courseOld)
                elif courseChoice.get() ==0:
                    courseExcel.set("")
            def get_existing_no():
                if noChoice.get()==1:
                    noOld = excel_data['F'+str(each_cell)].value
                    noExcel.set(noOld)
                elif noChoice.get() ==0:
                    noExcel.set("")
            def get_existing_address():
                if addressChoice.get()==1:
                    addressOld = excel_data['H'+str(each_cell)].value
                    addressExcel.set(addressOld)
                elif addressChoice.get() ==0:
                    addressExcel.set("")
            
            def update():
                excel_data['A'+str(each_cell)].value = student_noTxt.get()
                excel_data['B'+str(each_cell)].value = full_nameTxt.get()
                excel_data['C'+str(each_cell)].value = emailTxt.get()
                excel_data['D'+str(each_cell)].value = genderTxt.get()
                excel_data['E'+str(each_cell)].value = courseTxt.get()
                excel_data['F'+str(each_cell)].value = noTxt.get()
                excel_data['G'+str(each_cell)].value = addressTxt.get()

                excel_con.save('Sample_file.xlsx')
                messagebox.showinfo("UPDATED","DATA HAS BEEN UPDATED")
                Edit_form.destroy()
                search_box.delete(0, END)
                refresh_data(tv1)
            EditBtn = Button(Edit_form, width=15, font=("Arial", 15), text="Update Value",command=lambda:update(), bg="#b1f2ff")
            EditBtn.pack()
            EditBtn_w = my_canva.create_window(90,700, anchor="nw", window=EditBtn)

            Edit_form.mainloop()

    #Frames
    topframe = Frame(NewRoot, height=70, width=1030, bg="#CCAAFF")
    topframe.grid(row=0, column=0, columnspan=6, rowspan=2, sticky=W)
    centerframe = Frame(NewRoot, height=590, width=1030, bd=3)
    centerframe.grid(row=2, column=0, columnspan=3, rowspan=6, sticky=W)
    bottomframe = Frame(NewRoot, height=100, width=1030, bg="#CCAAFF")
    bottomframe.grid(row=8, column=0, columnspan=6, rowspan=2, sticky=W)

    # topframe
    pat = "back.png"
    b = Image.open(pat)
    resize_b = b.resize((1030, 100))
    b = ImageTk.PhotoImage(resize_b)
    my_canv = Canvas(topframe, height=70, width=1030)
    my_canv.grid(row=0, column=0)
    my_canv.create_image(0,0,image=b, anchor="nw")
    my_canv.create_text(515, 40, text="STUDENT RECORDS SYSTEM", font=("System", 41),fill="#FFB703")

    #center
    path = "back.png"
    bg = Image.open(path)
    resize_bg = bg.resize((1030, 590))
    bg = ImageTk.PhotoImage(resize_bg)
    my_canva = Canvas(centerframe, height=590, width=1030)
    my_canva.grid()
    my_canva.create_image(0,0,image=bg, anchor="nw")

    student_no = LabelFrame(centerframe, text="Student No.", font=("Arial", 10))
    studentno_Entry = Entry(student_no, width=30, font=("Arial", 15))
    student_no.grid(row=0, column=0)
    studentno_Entry.grid(row=0, column=1, padx=10, pady=10,)
    student_no_canva = my_canva.create_window(50, 30, anchor="nw", window=student_no)

    full_name = LabelFrame(centerframe, text="Full Name", font=("Arial", 10))
    fullname_Entry = Entry(full_name, width=30, font=("Arial", 15))
    full_name.grid(row=1, column=0)
    fullname_Entry.grid(row=1, column=1, padx=10, pady=10)
    full_name_canva = my_canva.create_window(50, 110, anchor="nw", window=full_name)

    email = LabelFrame(centerframe, text="Email", font=("Arial", 10))
    email_Entry = Entry(email, width=30, font=("Arial", 15))
    email.grid(row=2, column=0)
    email_Entry.grid(row=2, column=1, padx=10, pady=10)
    email_canva = my_canva.create_window(50, 190, anchor="nw", window=email)

    gender = LabelFrame(centerframe, text="Gender", font=("Arial", 10))
    gender_var = StringVar()
    gender_var.set("none")
    male_R = Radiobutton(gender, text="Male", variable=gender_var, value="Male", width=17, font=("Arial", 12))
    female_R = Radiobutton(gender, text="Female", variable=gender_var, value="Female", width=16, font=("Arial", 12))
    male_R.grid(row=3, column=0)
    female_R.grid(row=3, column=1)
    gender_canva = my_canva.create_window(50, 270, anchor="nw", window=gender)

    course = LabelFrame(centerframe, text="Course", font=("Arial", 10))
    course_var = StringVar()
    course_list = ["BSIT", "BSA", "BSAIS", "ABELS", "BSSW", "BSE", "DHRS", "BSPA", "BTVTE"]
    course_combo = ttk.Combobox(course, values=course_list, textvariable=course_var, font=("Arial", 15),width=30)
    course_combo.grid(row=4, column=0, padx=1, pady=10)
    course_canva = my_canva.create_window(50, 330, anchor="nw", window=course)

    no = LabelFrame(centerframe, text="Contact No.", font=("Arial", 10))
    no_Entry = Entry(no, width=30, font=("Arial", 15))
    no.grid(row=5, column=0)
    no_Entry.grid(row=5, column=1, padx=10, pady=10)
    no_canva = my_canva.create_window(50, 410, anchor="nw", window=no)

    address = LabelFrame(centerframe, text="Address", font=("Arial", 10))
    address_Entry = Entry(address, width=30, font=("Arial", 15))
    address.grid(row=6, column=0)
    address_Entry.grid(row=7, column=1, padx=10, pady=10)
    address_canva = my_canva.create_window(50, 490, anchor="nw", window=address)

    student_details = LabelFrame(centerframe, text="Student Details", font=("Arial", 15))
    student_detailsT = Text(student_details, width=51, font=("Arial", 15), height=8)
    student_details.grid(row=1, column=1, columnspan=2)
    student_detailsT.grid(row=2, column=1)
    student_details_canva = my_canva.create_window(420, 30, anchor="nw", window=student_details)

    firstclick = True
    def on_search_box_click(event):     
        search_box.delete(0, "end")


    search_boxf = LabelFrame(centerframe, text="Search, Delete, Edit", font=("Arial", 12))
    search_box = Entry(search_boxf, width=39, font=("Arial", 12))
    search_box.insert(0, "Enter the Student No.")
    search_box.bind('<FocusIn>', on_search_box_click)
    search_boxf.grid(row=1, column=0)
    search_box.grid(row=1, column=1, padx=10, pady=10)


    s = Button(search_boxf, text="Search", bd=0, font=("Arial", 12),width=7, command=lambda:search_data())
    s.grid(row=1, column=2)
    d = Button(search_boxf, text="Delete", bd=0, font=("Arial", 12),width=7, command=lambda:delete_data())
    d.grid(row=1, column=3)
    e = Button(search_boxf, text="Edit", bd=0, font=("Arial", 12), width=5, command=lambda:edit_data())
    e.grid(row=1, column=4)
    search_boxf_canva = my_canva.create_window(420, 250, anchor="nw", window=search_boxf)

    t = LabelFrame(centerframe, text="VIEW", font=("Arial", 15))
    global tv1
    tv1 = ttk.Treeview(t, show='headings')
    treescrolly = Scrollbar(t, orient="vertical", command=tv1.yview)
    treescrollx = Scrollbar(t, orient="horizontal", command=tv1.xview)
    tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
    treescrollx.pack(side ="bottom",fill ="x")
    treescrolly.pack(side ="right",fill="y")  

    tv1['columns'] = ("Student No.", "Fullname", "Email", "Gender", "Course", "Contact No.", "Address")
    tv1.column("#0", width=80, minwidth=25)
    tv1.column("Student No.", anchor=W, width=80)
    tv1.column("Fullname",  anchor=W, width=120)
    tv1.column("Email", anchor=W, width=120)
    tv1.column("Gender", anchor=W, width=50)
    tv1.column("Course", anchor=W, width=50)
    tv1.column("Contact No.", anchor=W, width=80)
    tv1.column("Address", anchor=W, width=120)

    tv1.heading("#0", text="Label", anchor=W)
    tv1.heading("Student No.", text="Student No.", anchor=W)
    tv1.heading("Fullname", text="Fullname", anchor=W)
    tv1.heading("Email", text="Email", anchor=W)
    tv1.heading("Gender", text="Gender", anchor=W)
    tv1.heading("Course", text="Course", anchor=W)
    tv1.heading("Contact No.", text="Contact No.", anchor=W)
    tv1.heading("Address", text="Address", anchor=W)

    for each_cell in range(2, (excel_data.max_row)+1):
        tv1.insert(parent='', index="end", text=str(each_cell),values=(excel_data['A'+str(each_cell)].value,excel_data['B'+str(each_cell)].value, excel_data['C'+str(each_cell)].value, excel_data['D'+str(each_cell)].value, excel_data['E'+str(each_cell)].value, excel_data['F'+str(each_cell)].value, excel_data['G'+str(each_cell)].value, excel_data['H'+str(each_cell)].value))
    tv1.pack(fill="both", expand=True)
    t.place(x=420, y=315, width=570, height=245)

    # bottomframe
    pa = "back.png"
    c = Image.open(pa)
    resize_c = c.resize((1040, 100))
    c = ImageTk.PhotoImage(resize_c)
    my_can = Canvas(bottomframe, height=100, width=1040)
    my_can.grid(row=0, column=0)
    my_can.create_image(0,0,image=c, anchor="nw")


    #BUTTONS
    add_btn = Button(bottomframe, text="Add Record", font=("System", 25), width=13, command=lambda:add_record())
    add_btn_w = my_can.create_window(10, 25, anchor="nw", window=add_btn)

    print_btn = Button(bottomframe, text="Print", font=("System", 25), width=13, command=lambda:print_records())
    print_btn_w = my_can.create_window(270, 25, anchor="nw", window=print_btn)

    reset_btn = Button(bottomframe, text="Reset", font=("System", 25), width=13, command=lambda:reset_fields())
    reset_btn_w = my_can.create_window(530, 25, anchor="nw", window=reset_btn)   

    exit_btn = Button(bottomframe, text="Exit", font=("System", 25), width=13, command=lambda:exit_interface())
    exit_btn_w = my_can.create_window(790, 25, anchor="nw", window=exit_btn)


    NewRoot.mainloop()