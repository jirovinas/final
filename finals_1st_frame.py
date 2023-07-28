from  tkinter import *
import tkinter as tk 
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk
from finals_2nd_frame import newinterface 

#FIRST INTERFACE
root = tk.Tk()
root.geometry("1120x600")
root.title("Login Interface")
root.resizable(False, False)

excel_con = Workbook()
excel_con = load_workbook("Sample_file.xlsx")
excel_data = excel_con['Data']
excel_old = excel_con["Old Students"]
excel_new = excel_con["New Students"]


def oldReg():
	newroot = Toplevel()
	newroot.geometry("400x800")
	newroot.title("Register Interface For Old Students") 
	newroot.resizable(False, False)

	path = "ro.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((400, 800))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(newroot, height=800, width=400)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")


	studentno = Entry(newroot,font=("Arial", 20), width=18)
	studentno.pack()
	studentno_w = my_canva.create_window(62, 343, anchor="nw", window=studentno)

	year_section = Entry(newroot,font=("Arial", 20), width=18)
	year_section.pack()
	year_section_w = my_canva.create_window(62, 435, anchor="nw", window=year_section)

	password = Entry(newroot,font=("Arial", 20), width=18)
	password.pack()
	password_w = my_canva.create_window(62, 527, anchor="nw", window=password)

	def oldlyreg():
		s = studentno.get()
		ys = year_section.get()
		ps = password.get()
		if s == "" or ys == "" or  ps == "":
			messagebox.showerror("Error", "Please Complete Your Details")
		else:
			excel_old.append((studentno.get(), password.get(), year_section.get()))
			messagebox.showinfo("Register Message", "Register Successful")
			excel_con.save("Sample_file.xlsx")
			newroot.withdraw()

	r_button = Button(newroot, text="Register", width=20, bg="lightblue", font=("Arial", 15) , command=lambda:oldlyreg())
	button = my_canva.create_window(80, 600, anchor="nw", window=r_button)

	newroot.mainloop()
	

def newReg():
	
	nroot = Toplevel()
	nroot.geometry("400x800")
	nroot.title("Register Interface For New Students") 
	nroot.resizable(False, False)

	path = "rn.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((400, 800))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(nroot, height=800, width=400)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")


	fulln = Entry(nroot,font=("Arial", 20), width=18)
	fulln.pack()
	fulln_w = my_canva.create_window(62, 343, anchor="nw", window=fulln)

	year_sec = Entry(nroot,font=("Arial", 20), width=18)
	year_sec.pack()
	year_sec_w = my_canva.create_window(62, 435, anchor="nw", window=year_sec)

	passw = Entry(nroot,font=("Arial", 20), width=18)
	passw.pack()
	passw_w = my_canva.create_window(62, 527, anchor="nw", window=passw)

	def newlyreg():
		f = fulln.get()
		p = passw.get()
		y = year_sec.get()
		if f == "" or  p == "" or  y == "":
			messagebox.showerror("Error", "Please Complete Your Details")
		else:
			excel_new.append((fulln.get(), passw.get(), year_sec.get()))
			messagebox.showinfo("Register Message", "Register Successful")
			excel_con.save("Sample_file.xlsx")
			nroot.withdraw()

	r_button1 = Button(nroot, text="Register", width=20, bg="lightblue", font=("Arial",15),command=lambda:newlyreg())
	r_button = my_canva.create_window(80, 600, anchor="nw", window=r_button1)

	nroot.mainloop()

#Register Function
def reg():
	if str_var.get() == "Old":
		oldReg()
	elif str_var.get() == "New":
		newReg()
	else:
		messagebox.showerror("REGISTER ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")
		
def newLog():
	groot = Toplevel()
	groot.geometry("400x600")
	groot.title("Login Interface For New Students") 
	groot.resizable(False, False)

	path = "ln.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((400, 600))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(groot, height=600, width=400)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	fullname = Entry(groot,font=("Arial", 18))
	fullname.pack()
	fullname_w = my_canva.create_window(68, 227, anchor="nw", window=fullname)

	passwor = Entry(groot,font=("Arial", 18), show="*")
	passwor.pack()
	password = my_canva.create_window(68, 295, anchor="nw", window=passwor)

	def getUsers(fullname, passwor):
		id = 1
		isExisted = False
		for data in excel_new.iter_rows(values_only=True):
			if data[0] == fullname and data[1] == passwor:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			groot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			groot.destroy()
			root.withdraw()
			newinterface()

	button1 = Button(groot, text="Login", width=20, bg="lightblue", font=("Arial", 15), command=lambda:getUsers(fullname.get(), passwor.get()))
	button1 = my_canva.create_window(85, 380, anchor="nw", window=button1)
	groot.mainloop()	


def oldLog():
	vroot = Toplevel()
	vroot.geometry("400x600")
	vroot.title("Login Interface For Old Students") 
	vroot.resizable(False, False)

	path = "lo.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((400, 600))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(vroot, height=600, width=400)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	studentn = Entry(vroot,font=("Arial", 18))
	studentn.pack()
	studentn_w = my_canva.create_window(68, 227, anchor="nw", window=studentn)

	passwo = Entry(vroot,font=("Arial", 18), show="*")
	passwo.pack()
	passwo_w = my_canva.create_window(68, 295, anchor="nw", window=passwo)

	def getUser(studentn, passwo):
		id = 1
		isExisted = False
		for data in excel_old.iter_rows(values_only=True):
			if data[0] == studentn and data[1] == passwo:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			vroot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			vroot.destroy()
			root.withdraw()
			newinterface()

	button1 = Button(vroot, text="Login", width=20, bg="lightblue", font=("Arial",15), command=lambda:getUser(studentn.get(), passwo.get()))
	button1 = my_canva.create_window(85, 380, anchor="nw", window=button1)

	vroot.mainloop()
#Login function	
def log():
	if str_var.get() == "Old":
		oldLog()
	elif str_var.get() == "New":
		newLog()
	else:
		messagebox.showerror("LOGIN ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")


#FRAMES
topframe = Frame(root, height=50, width=1120, bg="#023047")
topframe.pack(fill="x", side="top")
leftframe = Frame(root, height=600, width=350, bg="#023047", bd=0)
leftframe.pack(fill="y",side='left')
rigthframe = Frame(root, height=600, width=820)
rigthframe.pack(fill="both", side='right')

#TEXT IN TOPFRAME
l = Label(topframe,text="Dalubhasaan Ng Lungsod Ng Lucena", bg="#023047", fg="#FFB703", font=("Arial", 30), justify="center")
l.pack()

#LOGO IMAGE
b = (Image.open("dll_logo.png"))
resized_image = b.resize((190, 190))
new_image = ImageTk.PhotoImage(resized_image)
logo = Label(leftframe, image=new_image, bg="#023047")
logo.pack()

#RadioButton
rFrame = Frame(leftframe)

str_var = StringVar()
str_var.set("none")
old_R = Radiobutton(rFrame, text="Old", variable=str_var, value="Old", font=("Arial", 15),bg="#023047",fg="#FFB703")
new_R = Radiobutton(rFrame, text="New", variable=str_var, value="New", font=("Arial", 15),bg="#023047",fg="#FFB703")
old_R.grid(row=0, column=0)
new_R.grid(row=0, column=1)

rFrame.pack()

#Buttons
rb = Button(leftframe, text="Register", width=15, font=("Arial", 12), command=lambda: reg())
lb = Button(leftframe, text="Login", width=15, font=("Arial", 12), command=lambda: log())
rb.pack(pady=15)
lb.pack()

def animateAbout():
	global aboutIterate, txt
	if aboutIterate < len(a):
		txt += a[aboutIterate] + "\n"
		a_l1.config(text=txt)
		aboutIterate += 1
		a_l1.after(200, animateAbout)

#TEXT BUTTON 
def showProf():
	prof.pack(fill="both")
	abt.pack_forget()

def showAbt():
	abt.pack(fill="both")
	prof.pack_forget()
	animateAbout()

abt_b = Button(topframe, text="About", command=showAbt, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
abt_b.pack(side=RIGHT, anchor=N)
prof_b = Button(topframe, text="Profile", command=showProf, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
prof_b.pack(side=RIGHT, anchor=N)

#FRAME FOR THE PROFILE AND ABOUT
prof = Frame(rigthframe, bg="#023047")

#SCHOOL IMAGE
bg = (Image.open("school.jpg"))
resized_imag = bg.resize((1120, 600))
new_imag = ImageTk.PhotoImage(resized_imag)
my_canva = Canvas(prof, height=600, width=1080)
my_canva.pack()
my_canva.create_image(0,0,image=new_imag, anchor="nw")
prof.pack()

aboutIterate = 0
txt = ""

#About
_a = "Dalubhasaan ng Lungsod ng Lucena"
a = [
	"One of the high impact programs of Mayor Roderick A. Alcala is free quality tertiary education.",
	"When he assumed office in 2012, Dalubhasaan ng Lungsod ng Lucena (DLL) was his vision of providing access to college education for free. \nMayor Alcala envisions DLL as an institution that would provide easy access to higher education and \nultimately develop the competencies of the youth of the city to meet the demands of the local industries and businesses.",
	"Through DLL, students from low-income families are able to enrol in degree programs at no cost. \nThe annual appropriation of the local government has allowed DLL to cover all its operation expenses \nincluding tuition and miscellaneous fees of students. \nThe college, operated, managed, and fully-subsidized by the City Government, implements a zero-collection policy.",
	"At present, DLL has a total of nine degree programs and continue to apply for additional academic programs to accommodate more scholars:",
	"Bachelor of Arts in Information Technology",
	"Bachelor of Arts in Public Administration",
	"Bachelor of Science in Accountancy",
	"Bachelor of Science in Accounting Information System",
	"Bachelor of Science in English Language Studies",
	"Bachelor of Science in Entrepreneurship",
	"Bachelor of Science in Social Work",
	"Bachelor in Technical Vocational Teachers Education",
	"Diploma in Hotel and Restaurant Services",
	"At present the former annex building of Lucena City Hall is being renovated to accommodate the growing student population of DLL."
]

abt = Frame(rigthframe, width=600, height=750)
a_l = Label(abt,text=_a,font=("Arial" ,25),fg="black",justify="center")
a_l1 = Label(abt,font=("Arial" ,11),fg="black",justify="left")

a_l.pack()
a_l1.pack(pady=20)


abt.pack()
abt.mainloop()

showProf()


root.mainloop()

